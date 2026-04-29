"""报告生成器 — JSON、Markdown、Excel、HTML 格式输出，支持按规格分组。"""

import json
from datetime import datetime
from pathlib import Path

import plotly.graph_objects as go
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from price_monitor.analysis.trend import TrendResult
from price_monitor.analysis.value_rank import ValueScore


class ReportGenerator:
    @staticmethod
    def to_json(grouped: dict[str, list[ValueScore]]) -> str:
        data = {
            "generated_at": datetime.now().isoformat(),
            "groups": {
                key: [r.to_dict() for r in rankings]
                for key, rankings in grouped.items()
            },
        }
        return json.dumps(data, ensure_ascii=False, indent=2)

    @staticmethod
    def to_markdown(
        grouped: dict[str, list[ValueScore]],
        overall: list[ValueScore] | None = None,
    ) -> str:
        if not grouped:
            return "# 内存条性价比排名\n\n暂无数据。\n"

        lines = [
            "# 内存条规格对比",
            "",
            f"> 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"> 共 {len(grouped)} 个规格组",
            "",
        ]

        total_products = 0
        for group_key, rankings in grouped.items():
            total_products += len(rankings)
            lines.append(f"## {group_key}")
            lines.append("")
            lines.append("| 品牌 | 型号 | 类型 | 颗粒 | 时序 | 到手价 | ¥/GB | 建议 |")
            lines.append("|------|------|------|------|------|--------|------|------|")

            for r in rankings:
                price_yuan = f"¥{r.final_fen / 100:.2f}"
                ppg = f"¥{r.price_per_gb:.1f}"
                cl_str = f"CL{r.cl_latency}" if r.cl_latency else "-"
                die_str = r.die_type or "-"
                rec_emoji = {"buy_now": "🔥购入", "watch": "👀观望", "wait": "⏸等待", "accumulating": "📊积累中"}.get(
                    r.recommendation, "?"
                )
                lines.append(
                    f"| {r.brand} | {r.model[:30]} | {r.form_factor} | "
                    f"{die_str} | {cl_str} | {price_yuan} | {ppg} | {rec_emoji} |"
                )
            lines.append("")

        lines.append(f"> 共 {total_products} 款商品，{len(grouped)} 个规格组")
        return "\n".join(lines)

    @staticmethod
    def to_excel(grouped: dict[str, list[ValueScore]], filepath: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "内存性价比排名"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        data_font = Font(name="微软雅黑", size=10)
        rec_labels = {"buy_now": "🔥建议购入", "watch": "👀观望", "wait": "⏸等待", "accumulating": "📊数据积累中"}

        headers = ["规格", "品牌", "型号", "类型", "颗粒", "时序", "到手价(¥)", "¥/GB", "建议"]
        col_widths = [32, 14, 36, 8, 12, 10, 14, 8, 14]

        # Write headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Flatten all groups into rows
        row = 2
        for group_key, rankings in grouped.items():
            for r in rankings:
                cl_str = f"CL{r.cl_latency}" if r.cl_latency else "-"
                die_str = r.die_type or "-"
                price_yuan = round(r.final_fen / 100, 2)
                values = [
                    group_key,
                    r.brand,
                    r.model[:60],
                    r.form_factor,
                    die_str,
                    cl_str,
                    price_yuan,
                    round(r.price_per_gb, 1),
                    rec_labels.get(r.recommendation, "⏸等待"),
                ]

                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    if col in (6, 7, 8, 9):
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                row += 1

        # Column widths
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Freeze header
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{row - 1}"

        wb.save(filepath)

    @staticmethod
    def to_feishu_card(grouped: dict[str, list[ValueScore]], top_n: int = 3) -> str:
        """生成飞书富文本卡片 JSON。每个规格组展示 TOP N。"""
        if not grouped:
            return json.dumps({"msg_type": "text", "content": {"text": "暂无性价比数据"}})

        elements = [
            {
                "tag": "div",
                "text": {"tag": "lark_md", "content": "**🛒 内存条性价比速报（同规格对比）**"},
            },
            {"tag": "hr"},
        ]

        for group_key, rankings in grouped.items():
            elements.append({
                "tag": "div",
                "text": {"tag": "lark_md", "content": f"**▸ {group_key}**"},
            })

            for r in rankings[:top_n]:
                price_yuan = f"¥{r.final_fen / 100:.2f}"
                cl_str = f"CL{r.cl_latency}" if r.cl_latency else ""
                die_str = f" {r.die_type}" if r.die_type else ""
                rec = {"buy_now": "🔥建议购入", "watch": "👀观望", "wait": "⏸等待", "accumulating": "📊数据积累中"}.get(
                    r.recommendation, ""
                )
                elements.append({
                    "tag": "div",
                    "text": {
                        "tag": "lark_md",
                        "content": f"{r.brand} {r.model}{die_str} | {cl_str} | {price_yuan} | {rec}",
                    },
                })

            elements.append({"tag": "hr"})

        total = sum(len(v) for v in grouped.values())
        elements.append({
            "tag": "div",
            "text": {
                "tag": "lark_md",
                "content": f"共 {len(grouped)} 个规格组、{total} 款商品",
            },
        })

        return json.dumps(
            {"msg_type": "interactive", "card": {"elements": elements}},
            ensure_ascii=False,
        )

    @staticmethod
    def to_html(
        grouped: dict[str, list[ValueScore]],
        price_histories: dict[int, list[dict]],
        trends: list[TrendResult],
    ) -> str:
        """生成带 Plotly 交互图表的 HTML 报告。"""
        env = Environment(
            loader=FileSystemLoader(Path(__file__).parent.parent / "templates"),
            autoescape=False,
        )
        template = env.get_template("report.html")

        trend_map = {t.product_id: t for t in trends}

        # Compute group-level daily averages for overview chart
        overview_fig = ReportGenerator._build_overview_chart(grouped, price_histories)

        groups_data = []
        all_pids = []
        for group_key, products in grouped.items():
            group_pids = [p.product_id for p in products]
            all_pids.extend(group_pids)
            group_histories = {pid: h for pid, h in price_histories.items() if pid in group_pids}
            group_chart = ReportGenerator._build_group_chart(group_key, group_histories)

            products_data = []
            for p in products:
                trend = trend_map.get(p.product_id)
                rec = trend.recommendation if trend else "wait"
                data_days = trend.data_days if trend else 0
                if rec == "accumulating":
                    rec_label = f"📊积累中 {data_days}/7天"
                else:
                    rec_label = {"buy_now": "🔥建议购入", "watch": "👀观望", "wait": "⏸等待"}.get(rec, "?")
                products_data.append({
                    "product_id": p.product_id,
                    "brand": p.brand,
                    "model": p.model[:40],
                    "title": p.title,
                    "die_type": p.die_type,
                    "cl_str": f"CL{p.cl_latency}" if p.cl_latency else "-",
                    "fwl_str": f"{p.fwl_ns:.1f}ns" if p.fwl_ns else "-",
                    "perf_tier": p.performance_tier,
                    "price_yuan": f"{p.final_fen / 100:.2f}",
                    "price_per_gb": f"{p.price_per_gb:.1f}",
                    "form_factor": p.form_factor,
                    "brand_tier": p.brand_tier,
                    "trend_signal": trend.trend_signal if trend else "stable",
                    "trend_label": {"falling": "⬇下跌", "rising": "⬆上涨", "stable": "→持平"}.get(
                        trend.trend_signal if trend else "stable", "?"
                    ),
                    "recommendation": rec,
                    "rec_label": rec_label,
                })

            groups_data.append({
                "key": group_key,
                "count": len(products),
                "products": products_data,
                "chart_json": group_chart.to_json() if group_chart else None,
            })

        # Determine date range
        all_dates = set()
        for h in price_histories.values():
            for row in h:
                if row.get("bucket"):
                    all_dates.add(row["bucket"][:10])
        date_range = f"{min(all_dates)} → {max(all_dates)}" if all_dates else "暂无数据"

        return template.render(
            title="内存条价格趋势报告",
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
            stats={
                "total_products": sum(len(v) for v in grouped.values()),
                "total_groups": len(grouped),
                "date_range": date_range,
                "platforms": "京东",
            },
            overview_json=overview_fig.to_json() if overview_fig else None,
            groups=groups_data,
            product_histories_json=json.dumps(
                {str(pid): h for pid, h in price_histories.items()},
                ensure_ascii=False,
            ),
        )

    @staticmethod
    def _build_overview_chart(
        grouped: dict[str, list[ValueScore]],
        price_histories: dict[int, list[dict]],
    ) -> go.Figure | None:
        """构建规格组均价走势总览图。"""
        fig = go.Figure()

        colors = ["#00d4aa", "#4dc9f6", "#ffd93d", "#ff6b6b", "#c084fc",
                   "#fb923c", "#38bdf8", "#a3e635", "#f472b6", "#94a3b8"]

        for i, (group_key, products) in enumerate(grouped.items()):
            pids = {p.product_id for p in products}
            # Collect daily averages
            daily: dict[str, list[float]] = {}
            for pid in pids:
                history = price_histories.get(pid, [])
                for row in history:
                    if row.get("bucket") and row.get("avg_final_fen"):
                        date = row["bucket"][:10]
                        daily.setdefault(date, []).append(row["avg_final_fen"] / 100)
            if not daily:
                continue
            dates = sorted(daily.keys())
            avgs = [sum(daily[d]) / len(daily[d]) for d in dates]
            color = colors[i % len(colors)]
            fig.add_trace(go.Scatter(
                x=dates, y=avgs, mode="lines+markers",
                name=group_key,
                line={"color": color, "width": 2},
                marker={"size": 4},
            ))

        if len(fig.data) == 0:
            return None

        fig.update_layout(
            margin={"l": 48, "r": 16, "t": 8, "b": 32},
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font={"color": "#8b8fa3", "size": 11},
            xaxis={"gridcolor": "rgba(42,45,58,0.5)", "showgrid": True, "dtick": "D"},
            yaxis={"gridcolor": "rgba(42,45,58,0.5)", "showgrid": True, "title": "¥", "zeroline": False},
            legend={"orientation": "h", "y": 1.1, "font": {"size": 10}},
            hovermode="x unified",
        )
        return fig

    @staticmethod
    def _build_group_chart(
        group_key: str,
        group_histories: dict[int, list[dict]],
    ) -> go.Figure | None:
        """构建规格组均价走势 mini-chart。"""
        daily: dict[str, list[tuple[float, float]]] = {}  # date -> [(avg, min)]
        for history in group_histories.values():
            for row in history:
                if row.get("bucket") and row.get("avg_final_fen"):
                    date = row["bucket"][:10]
                    daily.setdefault(date, []).append(
                        (row["avg_final_fen"] / 100, row.get("min_final_fen", 0) / 100)
                    )
        if not daily:
            return None
        dates = sorted(daily.keys())
        avgs = [sum(p[0] for p in daily[d]) / len(daily[d]) for d in dates]
        mins = [min(p[1] for p in daily[d]) for d in dates]

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=dates, y=avgs, mode="lines+markers",
            name="均价", line={"color": "#00d4aa", "width": 2},
            marker={"size": 4},
        ))
        fig.add_trace(go.Scatter(
            x=dates, y=mins, mode="lines",
            name="最低价", line={"color": "#4dc9f6", "width": 1, "dash": "dot"},
        ))
        fig.update_layout(
            margin={"l": 48, "r": 16, "t": 8, "b": 32},
            height=200,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font={"color": "#8b8fa3", "size": 10},
            xaxis={"gridcolor": "rgba(42,45,58,0.5)", "showgrid": True, "dtick": "D"},
            yaxis={"gridcolor": "rgba(42,45,58,0.5)", "showgrid": True, "zeroline": False},
            showlegend=True,
            legend={"orientation": "h", "y": 1.15, "font": {"size": 9}},
            hovermode="x unified",
        )
        return fig
